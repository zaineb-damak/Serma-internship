import os
import pandas as pd
import json
from openpyxl import load_workbook
from Connected_Devices import *

#returns global station number corresponding to the given device id
#takes as parameter: phone id
def getGSNumber(phone_id):
    f = open('./application/GlobalStation.json')
    data = json.load(f)
    for i in data:
         if(i['DeviceID'] == phone_id):
             return i['GSNumber']
    return('wrong id')


#returns global station number corresponding to the given device name
#takes as parameter: phone name
def getGSNumberByName(phone_name):
    f = open('./application/GlobalStation.json')
    data = json.load(f)
    for i in data:
        if(i['DeviceName'] == phone_name):
            return i['GSNumber']
    return('wrong id')

#returns global station device name corresponding to the given device id
#takes as parameter: phone id
def getGSNameById(phone_id):
    f = open('./application/GlobalStation.json')
    data = json.load(f)
    for i in data:
            if(i['DeviceID'] == phone_id):
                return i['DeviceName']
    return('wrong id')

#returns global station device name corresponding to the given gsNumber
##takes as parameter: globale station phone number
def getGSNameByNumber(phone_number):
    f = open('./application/GlobalStation.json')
    data = json.load(f)
    for i in data:
            if(int(i['GSNumber']) == phone_number):
                return i['DeviceName']
    return('wrong id')

#returns a list of device names in global station 
def deviceNames():
    f = open('./application/GlobalStation.json')
    data = json.load(f)
    names=[]
    for i in data:
        names.append(i['DeviceName'])
    return names

#saves the changes in a new excel file
#takes as parameter: wb: excel workbook in which the changes will be saved; it is a copy of the original excel file
#                    row: the row number which will be modified 
#                    column: the column number which will be modified
#                    data: the data that will be stored inside the excel cell
#                    new_file: the name of the updated excel file
#                    save_path: the path in which the updated excel file will be saved
def newExcel(wb, row, column, data, new_file, save_path):
    ws = wb.active
    ws.cell(row=row, column=column).value = data
    wb.save(filename = save_path+'/'+new_file +'.xlsx')

#returns the row number of the columns' names
#takes as parameter: path of the excel file
def getHeader(file_path):
    df = pd.read_excel(file_path)
    header = "STEP"
    for index, row in df.iterrows():
        for name in enumerate(row):
            if pd.notna(name) and header in name:
                # we add 1 because dataframe considers the first line in an excel file as a header and not as a line in the dataframe
                return index+1

#returns the number of the column corresponding to the given name of the column
#takes as parameters: list of columns in the dataframe and name of the column
def getColumn(columns, column_name):
     for column in range (0,len(columns)):
        if columns[column] == column_name:
            return column+1

#saves chnages made inside the dataframe
#takes as parameter: df: dataframe
#                    list: list of id of connected devices
#                    row: row number in the for loop
#                    index: index in the for loop
def changeActionField(df,list,row,index):
    corresponding_phone.append([row['Action Field 3'],int(getGSNumber(list[0])),getGSNameById(list[0])])
    row['Action Field 3'] = getGSNumber(list[0])
    df.loc[index,'Action Field 3']  = row['Action Field 3'] 
    list.pop(0)

# changes the name of the device in the excel file
#takes as parameter: file_path: path of the excel file
#                    df: the dataframe
#                    row: row number in the for loop
#                    index: index in the for loop  
#                    excel_row: the list of the rows that will be modified in the excel file
#                    id_list: list of id of connected devices
def chengeDeviceName(file_path,df, row, index, excel_row,id_list):
    #store the list of device names in variable
    names = deviceNames()
    #check if the phone's name is in the names list
    if str(row['Action Field 3']).replace('"','') in  names:
        #add the row number to the list of excel rows that will be modified
        #we add 2 because there's a difference of 2 lines between the dataframe and excel file
        excel_row.append(index+getHeader(file_path)+2)
        #store the old phone name in variable old
        old = getGSNumberByName(str(row['Action Field 3']).replace('"',''))
        #initialize variable exists. if old is in corresponding_phone, exists is true, otherwise, exists is false
        #this is in case the first time a phone appears in the excel file by name and not by the gsNumber
        exists = False
        #check if old is in correspondingy_phone
        for i in range(0,len(corresponding_phone)):
            if (int(old) == corresponding_phone[i][0]):
                exists = True
                break
        if exists == True:
            row['Action Field 3'] = corresponding_phone[i][2]
            df.loc[index,'Action Field 3']  = row['Action Field 3']
        if exists == False:
            corresponding_phone.append([int(old),int(getGSNumber(id_list[0])),getGSNameById(id_list[0])])
            row['Action Field 3'] = getGSNameById(id_list[0])
            df.loc[index,'Action Field 3']  = row['Action Field 3']
            id_list.pop(0)  

#all modification will be made by this function
#takes as parameter: file_path: path of the excel file
#                    connectedID: string: return of adb command. example: "List of devices attached R59RA00NL7D device LMG900EMf7a2d5d5 device R58M36NV1GD device R58N91KCNYY device 215cf1f7 device"
#                    new_file: the name of the updated excel file
#                    save_path: the path in which the updated excel file will be saved
def updateTestCase(file_path, connectedID, new_file, save_file):
    global corresponding_phone
    global excel_row
    #create a new dataframe
    df = pd.read_excel(file_path, header=getHeader(file_path))
    #empty list linking the old phone number to the new phone number
    corresponding_phone = []
    #initialize variable which will be later used in newExcel(); stores the rows in which changes were made
    excel_row = []
    #initialize the list of Id
    id_list = idList(connectedID)
    #changing column action field into string
    df['Action Field 1 ( Tag Name/Request Name/ Picture Name)'] = df['Action Field 1 ( Tag Name/Request Name/ Picture Name)'].astype(str)
    #iterating through each row 
    for index,row in df.iterrows():
        #check if there is a phone being used
        if('Phone' in row['Action Field 1 ( Tag Name/Request Name/ Picture Name)']):
            #each iteration add to the list the index of the row in which we made changes 
            excel_row.append(index+58)
            #if Action Field 3 is empty we don't execute the rest of the instructions
            if pd.isnull(row['Action Field 3'])==True:
                continue
            #case 1: corresponding_phone is empty
            if (len(corresponding_phone) == 0 ):
                #add element to list [old gsNumber, new gsNumber, new deviceName]
                changeActionField(df,id_list,row,index)
            else:
                #check if the phone has already been changed and stored in corresponding_phone
                #if true: execute case 2
                #if false: execute case 3
                found_matching_pair = False
                for i in range (0,len(corresponding_phone)):
                    if(row['Action Field 3'] == int(corresponding_phone[i][0])):
                        found_matching_pair = True
                        break
                #case 2: phone number exists in corresponding_phone
                if (found_matching_pair == True):
                    #replace the old value with the new value in dataframe
                    row['Action Field 3'] = int(corresponding_phone[i][1])
                    df.loc[index,'Action Field 3'] = row['Action Field 3'] 
                #case 3: phone number doesn't exist in corresponding_phone
                else:
                    changeActionField(df,id_list,row,index)
        
        #changing the names of the devices
        chengeDeviceName(file_path,df, row, index, excel_row,id_list)
       
    #saving the changes in a new excel file
    wb = load_workbook(file_path)
    for j in range (0,len(excel_row)):
        newExcel(wb, excel_row[j] ,getColumn(df.columns, 'Action Field 3'),df.loc[excel_row[j]-getHeader(file_path)-2,'Action Field 3'], new_file, save_file)



#returns list of names of test cases in the execution plan
#takes as parameter: the path of the execution plan file
def getTestCases(execution_plan_file_path):
    df= pd.read_excel(execution_plan_file_path)
    file_name=[]
    for index, row in df.iterrows():
        file_name.append(row['Test Name'] + '.xlsx') 
    return(file_name)

#updates the test cases present in the execution plan
#takes as parameter: execution_plan_file_path: the path of the execution plan file
#                    test_case_path: the path of the test case files in the execution plan
#                    idlist: list of id of connected devices
#                    save_path: the path in which the updated excel files will be saved

def executionPlan(execution_plan_file_path, test_case_path,idList, save_path):
    global newList
    #initialize variable newList which will be used in getChangesExecutionPlan(). stores corresponding_phone of each test case in the execution plan
    newList=[]
    #stores the names of the test cases
    testCases = getTestCases(execution_plan_file_path)
    for test in testCases:
        try:
            #gets the name of the test case without the extension '.xlsx'
            test= os.path.splitext(test)[0]
            updateTestCase(test_case_path+'/'+test+'.xlsx', idList,test+'_updated', save_path )
            newList.append(corresponding_phone)
            pass
        except FileNotFoundError:
            pass

#returns a list which contains the modifications made in the test case. the message will be displayed in the gui
def getChangesTestCase():
    global message
    old=[]
    new=[]
    message=[]
    i = 0
    for ele in corresponding_phone:
        old.append(getGSNameByNumber(ele[0])) 
        new.append(getGSNameByNumber(ele[1]))
        message.append('old phone: '+old[i]+' is changed into: '+ new[i])
        i = i+1
    return message

#returns a list which contains the modifications made in each test case. the message will be displayed in the gui 
def getChangesExecutionPlan():
    global message
    old=[]
    new=[]
    message=[]
    i = 0
    j = 0
    for l in newList:
        j = j+1
        message.append('test case: '+str(j))
        for ele in l:
            old.append(getGSNameByNumber(ele[0])) 
            new.append(getGSNameByNumber(ele[1]))
            message.append('old phone: '+old[i]+' is changed into: '+ new[i])
            i = i+1
    return message




listID=("List of devices attached R59RA00NL7D device LMG900EMf7a2d5d5 device R58M36NV1GD device R58N91KCNYY device 215cf1f7 device")
#updateTestCase('E:/stage SERMA summer 2023/application/resources/test_case6.xlsx',listID,'sample','E:/stage SERMA summer 2023/Serma-internship/application')
#executionPlan('./application/resources/CAN.xlsx','./application/resources/',listID,'E:/stage SERMA summer 2023/Serma-internship/application')
#print(message)
#print(getTestCases('./resources/CAN.xlsx'))
#print(getHeader('E:/stage SERMA summer 2023/application/resources/test_case6.xlsx'))


